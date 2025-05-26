#!/usr/bin/env python3
import pandas as pd
import re
from bs4 import BeautifulSoup
import sys
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Try to load spaCy model
nlp = None
try:
    import spacy
    nlp = spacy.load("en_core_web_sm")
except:
    print("Warning: Could not load spaCy model. Some PII detection may be limited.")
    print("To install: pip install spacy && python -m spacy download en_core_web_sm")

POSTAL_CODE_PATTERN = r"[A-Za-z]\s*\d\s*[A-Za-z]\s*[ -]?\s*\d\s*[A-Za-z]\s*\d"
PASSPORT_PATTERN = r"\b([A-Za-z]{2}\s*\d{6})\b"
SIN_PATTERN = r"(\d{3}\s*\d{3}\s*\d{3}|\d{3}\D*\d{3}\D*\d{3})"
PHONE_PATTERN_1 = r"(\+\d{1,2}\s?)?1?\-?\.\s?\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}"
PHONE_PATTERN_2 = r"(?:(?:\+?1\s*(?:[.-]\s*)?)?(?:\(\s*([2-9]1[02-9]|[2-9][02-8]1|[2-9][02-8][02-9])\s*\)|([2-9]1[02-9]|[2-9][02-8]1|[2-9][02-8][02-9]))\s*(?:[.-]\s*)?)?([2-9]1[02-9]|[2-9][02-9]1|[2-9][02-9]{2})\s*(?:[.-]\s*)?([0-9]{4})(?:\s*(?:#|x\.?|ext\.?|extension)\s*(\d+))?"
EMAIL_PATTERN = r"([a-zA-Z0-9_\-\.]+)\s*@([\sa-zA-Z0-9_\-\.]+)[\.\,]([a-zA-Z]{1,5})"
# Regex for detecting common Date of Birth formats
DOB_PATTERN = r"\b(\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}|\d{4}[-/.]\d{1,2}[-/.]\d{1,2})\b"
# Regex for detecting PO Box and General Delivery addresses
PO_BOX_PATTERN = r"\b(?:P(?:ost(?:al)?)?\.?\s*O(?:ffice)?\.?\s*Box|General Delivery)\s+\d+\b"
# Regex for detecting common street address formats
ADDRESS_PATTERN = r"\b[A-Za-z0-9]+\s+[A-Za-z0-9\s.-]+\s+(?:Street|St|Avenue|Ave|Road|Rd|Lane|Ln|Drive|Dr|Boulevard|Blvd|Crescent|Cres|Court|Ct)\b"
# Pattern to temporarily identify credit card numbers to prevent their accidental scrubbing.
TEMP_CREDIT_CARD_IGNORE_PATTERN = r"\b(?:\d{4}[ -]?\d{4}[ -]?\d{4}[ -]?\d{1,4}|\d{13,16})\b"

def clean_html(text):
    """Remove HTML tags from text"""
    if pd.isna(text):
        return text
    soup = BeautifulSoup(str(text), 'html.parser')
    return soup.get_text()

def scrub_pii(text):
    """Remove personally identifiable information from text"""
    if pd.isna(text):
        return text
    
    text = str(text)

    # BEGIN CC IGNORE/RESTORE LOGIC
    cc_matches = []
    def cc_match_replacer(match):
        cc_matches.append(match.group(0))
        # Using .format() for Python 3.8 f-string equivalent for this specific placeholder
        return "@@TEMP_CC_PLACEHOLDER_{}@@".format(len(cc_matches)-1)

    text = re.sub(TEMP_CREDIT_CARD_IGNORE_PATTERN, cc_match_replacer, text, flags=re.IGNORECASE)
    # END CC IGNORE LOGIC (part 1)
    
    # Replace patterns with ***
    text = re.sub(POSTAL_CODE_PATTERN, "***", text, flags=re.IGNORECASE)
    text = re.sub(PASSPORT_PATTERN, "***", text, flags=re.IGNORECASE)
    text = re.sub(SIN_PATTERN, "***", text, flags=re.IGNORECASE)
    text = re.sub(PHONE_PATTERN_1, "***", text, flags=re.IGNORECASE)
    text = re.sub(PHONE_PATTERN_2, "***", text, flags=re.IGNORECASE)
    text = re.sub(EMAIL_PATTERN, "***", text, flags=re.IGNORECASE)
    text = re.sub(DOB_PATTERN, "***", text, flags=re.IGNORECASE)
    text = re.sub(PO_BOX_PATTERN, "***", text, flags=re.IGNORECASE) # Added PO Box scrubbing
    text = re.sub(ADDRESS_PATTERN, "***", text, flags=re.IGNORECASE) # Modified Address scrubbing
    
    # Use spaCy to detect and replace person names, organizations, and locations
    if nlp:
        try:
            doc = nlp(text)
            # BEGIN UPDATED spaCy entity processing logic
            entities_to_scrub_tuples = []
            for ent in doc.ents:
                if ent.text.upper() in ["PII", "DOB"] and ent.label_ in ["ORG", "PERSON", "PRODUCT", "WORK_OF_ART", "LAW", "EVENT"]: # Add more common mis-labels if observed
                    # Skip these specific acronyms if they are misidentified
                    # print("Skipping false positive entity: {} ({})".format(ent.text, ent.label_)) # Optional: for debugging
                    pass
                elif ent.label_ in ["PERSON", "ORG", "GPE", "LOC"]:
                    entities_to_scrub_tuples.append((ent.start_char, ent.end_char, ent.label_))
            
            entities_to_scrub_tuples.sort(reverse=True, key=lambda x: x[0])
            
            for start, end, label in entities_to_scrub_tuples:
                text = text[:start] + "***" + text[end:]
            # END UPDATED spaCy entity processing logic
        except Exception as e: # Catch all exceptions from spaCy processing
            # print("SpaCy processing error: {}".format(e)) # Optional: for debugging
            pass
    
    # BEGIN CC RESTORE LOGIC (part 2)
    idx = 0
    while True: # Loop as long as placeholders are found
        # Using .format() for Python 3.8 f-string equivalent for this specific placeholder
        placeholder_to_find = "@@TEMP_CC_PLACEHOLDER_{}@@".format(idx)
        if placeholder_to_find in text:
            if idx < len(cc_matches): # Check if we have a corresponding original CC
                text = text.replace(placeholder_to_find, cc_matches[idx], 1) # Replace one instance
            else:
                # This case should ideally not happen if logic is correct
                text = text.replace(placeholder_to_find, "[UNEXPECTED_CC_PLACEHOLDER]", 1) # Avoid using *** for this
            idx += 1
        else:
            break # No more placeholders with the current idx, exit loop
    # END CC RESTORE LOGIC

    return text

def generate_session_colors(df):
    """Generate color codes for duplicate session IDs"""
    session_counts = df['SessionId'].value_counts()
    duplicate_sessions = session_counts[session_counts > 1].index
    
    # Generate colors for duplicates
    colors = ['#FFB6C1', '#98FB98', '#87CEEB', '#DDA0DD', '#F0E68C', 
              '#FFA07A', '#20B2AA', '#B0C4DE', '#FAFAD2', '#D8BFD8']
    
    session_colors = {}
    for i, session in enumerate(duplicate_sessions):
        session_colors[session] = colors[i % len(colors)]
    
    return session_colors

def process_excel_file(input_file, output_file):
    """Process the Excel file with all required transformations"""
    print("Reading file: {}".format(input_file))
    df = pd.read_excel(input_file)
    
    print("Processing data...")
    
    # Clean HTML from RawAnswer.Answer column
    if 'RawAnswer.Answer' in df.columns:
        print("Cleaning HTML tags from RawAnswer.Answer column...")
        df['RawAnswer.Answer_Cleaned'] = df['RawAnswer.Answer'].apply(clean_html)
    
    # Scrub PII from UserQuestion column only
    if 'UserQuestion' in df.columns:
        print("Scrubbing PII from UserQuestion column...")
        df['UserQuestion_Scrubbed'] = df['UserQuestion'].apply(scrub_pii)
    
    # Calculate userSatisfactionIndicator statistics (lowercase 'u')
    summary_data = {}
    if 'userSatisfactionIndicator' in df.columns:
        print("\nCalculating userSatisfactionIndicator statistics...")
        satisfaction_counts = df['userSatisfactionIndicator'].value_counts()
        total = len(df[df['userSatisfactionIndicator'].notna()])
        
        print("\n=== UserSatisfactionIndicator Statistics ===")
        for indicator, count in satisfaction_counts.items():
            percentage = (count / total) * 100 if total > 0 else 0
            print("{}: {} ({:.2f}%)".format(indicator, count, percentage))
        
        # Separate up/down statistics
        up_count = df[df['userSatisfactionIndicator'].str.contains('up', case=False, na=False)].shape[0]
        down_count = df[df['userSatisfactionIndicator'].str.contains('down', case=False, na=False)].shape[0]
        
        if total > 0:
            up_percentage = (up_count/total)*100
            down_percentage = (down_count/total)*100
            print("\nTotal UP: {} ({:.2f}%)".format(up_count, up_percentage))
            print("Total DOWN: {} ({:.2f}%)".format(down_count, down_percentage))
            
            # Store summary data for the summary sheet
            summary_data = {
                'Metric': ['Total Responses', 'Total UP', 'Total DOWN', 'UP Percentage', 'DOWN Percentage'],
                'Value': [total, up_count, down_count, "{:.2f}%".format(up_percentage), "{:.2f}%".format(down_percentage)]
            }
        print("=" * 40)
    
    # Generate session colors for duplicates
    session_colors = generate_session_colors(df)
    
    # Write to Excel with formatting
    print("\nWriting processed data to: {}".format(output_file))
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Processed Data', index=False)
        
        # Add summary sheet if we have satisfaction data
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Satisfaction Summary', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Processed Data']
        
        # Apply color coding to duplicate SessionIds
        from openpyxl.styles import PatternFill
        
        # Find SessionId column
        session_col = None
        for idx, col in enumerate(df.columns, 1):
            if col == 'SessionId':
                session_col = idx
                break
        
        if session_col:
            for row_idx, session_id in enumerate(df['SessionId'], 2):  # Start from row 2 (after header)
                if session_id in session_colors:
                    cell = worksheet.cell(row=row_idx, column=session_col)
                    fill = PatternFill(start_color=session_colors[session_id].replace('#', ''),
                                     end_color=session_colors[session_id].replace('#', ''),
                                     fill_type='solid')
                    cell.fill = fill
        
        # Auto-adjust column widths for all sheets
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print("\nProcessing complete! Output saved to: {}".format(output_file))

if __name__ == "__main__":
    input_file = "test2.xlsx"
    output_file = "processed_chatlog_{}.xlsx".format(datetime.now().strftime('%Y%m%d_%H%M%S'))
    
    try:
        process_excel_file(input_file, output_file)
    except Exception as e:
        print("Error processing file: {}".format(e))
        sys.exit(1)